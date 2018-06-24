# -*- coding: utf-8 -*-
"""
Created on Wed Jun 20 08:56:08 2018

@author: DP070009
"""
import sys
import os
from typing import Dict, Any

import win32com.client as win32
import pythoncom
from openpyxl import load_workbook


wb = load_workbook('eCOS_table_population_guide_completed v1.3_test.xlsx')
ws = wb['systables']


def create_dict_of_header_names(ws: openpyxl.worksheet.worksheet.Worksheet):
    dict_of_headers: Dict[String, int] = {}
    for index, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        for cell in col:
            dict_of_headers[cell.value] = index
    return dict_of_headers


def check_for_filtered_sheets(ws: openpyxl.worksheet.worksheet.Worksheet):
    headers = create_dict_of_header_names(ws)
    sheets_to_remove = []
    headers = create_dict_of_header_names(ws)
    for row in ws.iter_rows(min_col=headers['NAME'], max_col=headers['NAME'], min_row=ws.min_row + 1):
        if not ws.row_dimensions[row[0].row].hidden:
            for cell in row:
                sheets_to_remove.append((cell.value).lower())
    return sheets_to_remove


sheets_to_remove = check_for_filtered_sheets(ws)
file = 'eCOS_table_population_guide_completed v1.3_test1.xlsx'
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = 1
excel.DisplayAlerts = False
file = os.path.join(os.getcwd(), file)
wb = excel.Workbooks.Open(file)

i = 1
while i < wb.Sheets.Count:
    if wb.Sheets(i).Name in sheets_to_remove:
        wb.Sheets(i).Delete()
    else:
        i += 1

excel.DisplayAlerts = True
wb.Close(True)
excel.Application.Quit()
# pythoncom.CoUninitialize()
del excel
